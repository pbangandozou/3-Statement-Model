import os
import sys
import time
import math
import json
import argparse
import requests
import pandas as pd
import numpy as np
from statistics import mean
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ==========================
# Settings
# ==========================
USER_AGENT = "ThreeStatementModel/1.0 (prustide@gmail.com)"  # contact email or URL (required by SEC)
RATE_SLEEP = 0.25

session = requests.Session()
session.headers.clear()
session.headers.update({
    "User-Agent": USER_AGENT,
    "Accept-Encoding": "gzip, deflate",
    "Accept": "application/json, text/plain, */*"
})

def _get_json(url, timeout=20, retries=3, backoff=0.6):
    last_err = None
    for i in range(retries):
        try:
            r = session.get(url, timeout=timeout)
            if r.status_code == 200:
                return r.json()
            if r.status_code in (429, 500, 502, 503, 504):
                last_err = ValueError(f"HTTP {r.status_code} for {url}")
                time.sleep(backoff * (2 ** i))
                continue
            raise ValueError(f"HTTP {r.status_code} for {url}")
        except Exception as e:
            last_err = e
            time.sleep(backoff * (2 ** i))
    raise ValueError(last_err)

# ==========================
# Ticker / Name resolution
# ==========================
def _load_sec_ticker_maps():
    """
    Load SEC mapping from both endpoints and normalize into a single DataFrame:
      - https://www.sec.gov/files/company_tickers.json      -> dict of {idx: {cik_str, ticker, title}}
      - https://www.sec.gov/files/company_tickers_exchange.json -> LIST of {cik, ticker, title, exchange}
    Returns columns: ticker, cik (10-digit, zero-padded), title
    """
    def _fetch_json(url):
        data = _get_json(url)
        time.sleep(RATE_SLEEP)
        return data

    data1 = _fetch_json("https://www.sec.gov/files/company_tickers.json")           # dict
    try:
        data2 = _fetch_json("https://www.sec.gov/files/company_tickers_exchange.json")  # list
    except Exception:
        data2 = []

    rows = []

    # company_tickers.json (DICT -> values())
    if isinstance(data1, dict):
        for v in data1.values():
            if not isinstance(v, dict):
                continue
            ticker = (v.get("ticker") or "").strip()
            cik_str = v.get("cik_str")
            title = (v.get("title") or "").strip()
            if ticker and cik_str is not None:
                rows.append({
                    "ticker": ticker.upper(),
                    "cik": str(int(cik_str)).zfill(10),
                    "title": title
                })

    # company_tickers_exchange.json (LIST of dicts)
    if isinstance(data2, list):
        for v in data2:
            if not isinstance(v, dict):
                continue
            ticker = (v.get("ticker") or "").strip()
            # Some entries use "cik" instead of "cik_str"
            cik_any = v.get("cik_str", v.get("cik"))
            title = (v.get("title") or "").strip()
            if ticker and cik_any is not None:
                rows.append({
                    "ticker": ticker.upper(),
                    "cik": str(int(cik_any)).zfill(10),
                    "title": title
                })

    if not rows:
        raise ValueError("Could not load SEC ticker maps (both endpoints empty).")

    df = pd.DataFrame(rows).drop_duplicates(subset=["ticker"]).reset_index(drop=True)
    return df


def resolve_ticker_or_name(query: str) -> str:
    """
    Accepts a ticker (e.g., 'NVDA') or company name (e.g., 'NVIDIA') and returns the ticker.
    """
    m = _load_sec_ticker_maps()
    q = (query or "").strip()
    if not q:
        raise ValueError("Empty query for ticker/company name.")

    q_upper = q.upper()

    # 1) Exact ticker match
    exact = m[m["ticker"] == q_upper]
    if not exact.empty:
        return exact.iloc[0]["ticker"]

    # 2) Title contains all words
    words = [w for w in q_upper.split() if w]
    if words:
        filt = m["title"].fillna("").str.upper()
        mask = pd.Series(True, index=filt.index)
        for w in words:
            mask &= filt.str.contains(w, na=False)
        cand = m[mask]
        if not cand.empty:
            # Prefer the shortest title match as a heuristic
            cand = cand.assign(tlen=cand["title"].str.len())
            cand = cand.sort_values(["tlen", "ticker"])
            return cand.iloc[0]["ticker"]

    # 3) Title prefix match
    cand2 = m[m["title"].fillna("").str.upper().str.startswith(q_upper)]
    if not cand2.empty:
        cand2 = cand2.sort_values(["title", "ticker"])
        return cand2.iloc[0]["ticker"]

    raise ValueError(f"Could not resolve a ticker for '{query}'. Try the exact ticker.")

def get_cik(ticker: str) -> str:
    """
    Return the 10-digit zero-padded CIK for a given ticker.
    Uses the already-normalized SEC maps from _load_sec_ticker_maps().
    """
    m = _load_sec_ticker_maps()
    t = ticker.strip().upper()
    row = m[m["ticker"] == t]
    if not row.empty:
        return str(row.iloc[0]["cik"]).zfill(10)
    raise ValueError(f"Could not find CIK for ticker {ticker}")

# ==========================
# XBRL helpers
# ==========================
TAGS = {
    "Revenue": ["Revenues", "SalesRevenueNet", "RevenueFromContractWithCustomerExcludingAssessedTax"],
    "COGS": ["CostOfGoodsAndServicesSold", "CostOfSales"],
    "Operating_Expenses": ["OperatingExpenses"],
    "Depreciation": ["DepreciationDepletionAndAmortization", "DepreciationAndAmortization"],
    "Interest_Expense": ["InterestExpense", "InterestAndDebtExpense"],
    "Net_Income": ["NetIncomeLoss"],
    "Tax_Expense": ["IncomeTaxExpenseBenefit", "IncomeTaxExpenseBenefitContinuingOperations"],
    "Accounts_Receivable": ["AccountsReceivableNetCurrent", "ReceivablesNetCurrent"],
    "Inventory": ["InventoryNet"],
    "Accounts_Payable": ["AccountsPayableCurrent"],
    "PP_and_E": ["PropertyPlantAndEquipmentNet"],
    "Cash": ["CashAndCashEquivalentsAtCarryingValue"],
    "DebtCurrent": ["DebtCurrent"],
    "DebtLT": ["LongTermDebtNoncurrent", "LongTermDebt"],
    "Equity": ["StockholdersEquity", "StockholdersEquityIncludingPortionAttributableToNoncontrollingInterest"],
    "Retained_Earnings": ["RetainedEarningsAccumulatedDeficit"]
}

def _company_facts_url(cik):
    return f"https://data.sec.gov/api/xbrl/companyfacts/CIK{cik}.json"

def _norm_fy(v):
    try: return int(v)
    except: return -10**9

def _norm_end(v): return v or ""

def _pick_annual_value(series, year):
    ordered = sorted(series, key=lambda r: (_norm_fy(r.get("fy")), _norm_end(r.get("end"))), reverse=True)
    for row in ordered:
        if str(row.get("fy")) == str(year) and row.get("fp") in ("FY", "FYR"):
            try: return float(row.get("val"))
            except: continue
    for row in ordered:
        end = str(row.get("end", ""))
        if str(year) in end:
            try: return float(row.get("val"))
            except: continue
    return None

def _get_fact_year(facts_json, tag_names, year):
    usgaap = facts_json.get("facts", {}).get("us-gaap", {})
    for tag in tag_names:
        node = usgaap.get(tag)
        if not node: continue
        units = node.get("units", {})
        series = units.get("USD") or (units[list(units.keys())[0]] if units else None)
        if series:
            val = _pick_annual_value(series, year)
            if val is not None:
                return val
    return None

def _safe(val, default=0.0):
    return default if val is None or (isinstance(val, float) and (math.isnan(val) or math.isinf(val))) else val

# ==========================
# Build 2020–2024 Actuals
# ==========================
def _build_historical_income(facts, years):
    rows = []
    for y in years:
        rev = _safe(_get_fact_year(facts, TAGS["Revenue"], y), 0.0)
        cogs = _safe(_get_fact_year(facts, TAGS["COGS"], y), 0.0)
        opex = _safe(_get_fact_year(facts, TAGS["Operating_Expenses"], y), 0.0)
        d_and_a = _safe(_get_fact_year(facts, TAGS["Depreciation"], y), 0.0)
        int_exp = _safe(_get_fact_year(facts, TAGS["Interest_Expense"], y), 0.0)
        net_inc = _get_fact_year(facts, TAGS["Net_Income"], y)
        tax_exp = _get_fact_year(facts, TAGS["Tax_Expense"], y)

        gross = rev - cogs
        ebitda = gross - opex
        ebit = ebitda - d_and_a
        ebt = ebit - int_exp

        if tax_exp is None and net_inc is not None:
            tax_exp = max(0.0, ebt - net_inc)

        tax_rate = 0.21
        if tax_exp is not None and ebt != 0:
            tax_rate = max(0.0, min(0.40, tax_exp / ebt))

        if net_inc is None:
            taxes = max(0.0, ebt * tax_rate)
            net_inc = ebt - taxes

        rows.append({
            "Year": int(y),
            "Revenue": rev,
            "COGS": cogs,
            "Gross_Profit": gross,
            "Operating_Expenses": opex,
            "EBITDA": ebitda,
            "Depreciation": d_and_a,
            "EBIT": ebit,
            "Interest_Expense": int_exp,
            "EBT": ebt,
            "Taxes": max(0.0, ebt * tax_rate),
            "Net_Income": _safe(net_inc, 0.0),
            "Tax_Rate": tax_rate
        })
    return pd.DataFrame(rows)

def _build_historical_bs_cf(facts, years, income_df):
    bs_rows = []
    cf_rows = []

    last_cash = last_ar = last_inv = last_ap = 0.0
    last_ppne = last_debt = last_equity = last_re = 0.0

    for idx, y in enumerate(years):
        y = int(y)
        cash = _safe(_get_fact_year(facts, TAGS["Cash"], y), last_cash)
        ar = _safe(_get_fact_year(facts, TAGS["Accounts_Receivable"], y), last_ar)
        inv = _safe(_get_fact_year(facts, TAGS["Inventory"], y), last_inv)
        ap = _safe(_get_fact_year(facts, TAGS["Accounts_Payable"], y), last_ap)
        ppne = _safe(_get_fact_year(facts, TAGS["PP_and_E"], y), last_ppne)
        debt = _safe(_get_fact_year(facts, TAGS["DebtCurrent"], y), 0.0) + _safe(_get_fact_year(facts, TAGS["DebtLT"], y), 0.0)
        equity = _safe(_get_fact_year(facts, TAGS["Equity"], y), last_equity)
        re = _safe(_get_fact_year(facts, TAGS["Retained_Earnings"], y), last_re)

        total_assets = cash + ar + inv + ppne
        total_liab = ap + debt
        total_equity = equity + re
        tle = total_liab + total_equity

        inc = income_df[income_df["Year"] == y].iloc[0]
        net_income = inc["Net_Income"]
        d_and_a = inc["Depreciation"]

        d_ar = -(ar - last_ar) if idx > 0 else 0.0
        d_inv = -(inv - last_inv) if idx > 0 else 0.0
        d_ap = (ap - last_ap) if idx > 0 else 0.0

        cfo = net_income + d_and_a + d_ar + d_inv + d_ap

        capex = max(0.0, (ppne - last_ppne) + d_and_a) if idx > 0 else 0.0
        cfi = -capex

        debt_issuance = max(0.0, debt - last_debt) if idx > 0 else 0.0
        debt_repay = -max(0.0, last_debt - debt) if idx > 0 else 0.0
        dividends = -max(0.0, 0.20 * net_income)
        stock_issuance = 0.0
        stock_buyback = 0.0
        cff = debt_issuance + debt_repay + stock_issuance + stock_buyback + dividends

        beginning_cash = cash if idx == 0 else last_cash
        net_change = cfo + cfi + cff
        ending_cash = cash  # align to reported

        bs_rows.append({
            "Year": y,
            "Cash": cash,
            "Accounts_Receivable": ar,
            "Inventory": inv,
            "PP_and_E": ppne,
            "Intangibles": 0.0,
            "Accounts_Payable": ap,
            "Debt": debt,
            "Equity": equity,
            "Retained_Earnings": re,
            "Total_Assets": total_assets,
            "Total_Liabilities": total_liab,
            "Total_Equity": total_equity,
            "Total_Liabilities_Equity": tle
        })

        cf_rows.append({
            "Year": y,
            "Net_Income": net_income,
            "Depreciation": d_and_a,
            "Change_in_AR": d_ar,
            "Change_in_Inventory": d_inv,
            "Change_in_AP": d_ap,
            "Cash_from_Operations": cfo,
            "CapEx": -capex,
            "Cash_from_Investing": -capex,
            "Debt_Issuance": debt_issuance,
            "Debt_Repayment": debt_repay,
            "Stock_Issuance": stock_issuance,
            "Stock_Buyback": stock_buyback,
            "Dividends": dividends,
            "Cash_from_Financing": cff,
            "Net_Change_in_Cash": net_change,
            "Beginning_Cash": beginning_cash,
            "Ending_Cash": ending_cash
        })

        last_cash, last_ar, last_inv, last_ap = cash, ar, inv, ap
        last_ppne, last_debt, last_equity, last_re = ppne, debt, equity, re

    return pd.DataFrame(bs_rows), pd.DataFrame(cf_rows)

# ==========================
# Assumptions (defaults + infer from history)
# ==========================
def infer_default_assumptions(income_hist, balance_hist, cashflow_hist):
    yrs = sorted(income_hist["Year"].astype(int).tolist())
    # growth
    revs = income_hist.set_index("Year")["Revenue"].sort_index()
    yoy = [(revs.iloc[i] / revs.iloc[i-1] - 1.0) for i in range(1, len(revs)) if revs.iloc[i-1] > 0]
    growth = (mean(yoy[-3:]) if len(yoy) >= 3 else (mean(yoy) if yoy else 0.05))

    def avg_ratio(nume, den):
        vals = []
        for y in yrs:
            d = den.get(y, 0)
            if d:
                vals.append(nume.get(y, 0) / d)
        return max(0.0, mean(vals)) if vals else 0.0

    inc_map = income_hist.set_index("Year").to_dict(orient="index")
    bs_map = balance_hist.set_index("Year").to_dict(orient="index")
    cf_map = cashflow_hist.set_index("Year").to_dict(orient="index")

    rev_dict = {y: inc_map[y]["Revenue"] for y in yrs}
    assumptions = {
        "Revenue_Growth_YoY": round(growth, 4),
        "COGS_pct_of_Revenue": round(avg_ratio({y: inc_map[y]["COGS"] for y in yrs}, rev_dict), 4),
        "OpEx_pct_of_Revenue": round(avg_ratio({y: inc_map[y]["Operating_Expenses"] for y in yrs}, rev_dict), 4),
        "DA_pct_of_Revenue": round(avg_ratio({y: inc_map[y]["Depreciation"] for y in yrs}, rev_dict), 4),
        "Interest_pct_of_Revenue": round(avg_ratio({y: inc_map[y]["Interest_Expense"] for y in yrs}, rev_dict), 4),
        "Tax_Rate": round(max(0.0, min(0.30, mean([inc_map[y]["Tax_Rate"] for y in yrs]))), 4) if yrs else 0.21,
        "AR_pct_of_Revenue": round(avg_ratio({y: bs_map[y]["Accounts_Receivable"] for y in yrs}, rev_dict), 4),
        "Inventory_pct_of_Revenue": round(avg_ratio({y: bs_map[y]["Inventory"] for y in yrs}, rev_dict), 4),
        "AP_pct_of_Revenue": round(avg_ratio({y: bs_map[y]["Accounts_Payable"] for y in yrs}, rev_dict), 4),
        "CapEx_pct_of_Revenue": round(avg_ratio({y: abs(cf_map[y]["CapEx"]) for y in yrs}, rev_dict), 4),
        # Debt drivers:
        "Interest_Rate_on_Debt": None,  # if None, infer from history
        "Debt_Issuance_pct_of_Revenue": 0.00,
        "Debt_Repayment_pct_of_Revenue": 0.00,
        "Dividend_Payout_pct_of_NI": 0.20
    }

    # Try infer interest rate from last 3 yrs: interest expense / avg debt
    try:
        last3 = yrs[-3:] if len(yrs) >= 3 else yrs
        rates = []
        for y in last3:
            debt = bs_map[y]["Debt"]
            if debt > 0:
                rates.append(inc_map[y]["Interest_Expense"] / debt)
        if rates:
            assumptions["Interest_Rate_on_Debt"] = round(max(0.0, mean(rates)), 4)
    except Exception:
        pass

    if assumptions["Interest_Rate_on_Debt"] is None:
        assumptions["Interest_Rate_on_Debt"] = 0.04  # safe default 4%

    return assumptions

# ==========================
# Projections (use assumptions + debt schedule)
# ==========================
def project_forward_with_assumptions(income_hist, balance_hist, cashflow_hist, assumptions, start_year=2025, n_years=5):
    hist_years = sorted(income_hist["Year"].astype(int).tolist())
    inc_map = income_hist.set_index("Year").to_dict(orient="index")
    bs_map = balance_hist.set_index("Year").to_dict(orient="index")
    cf_map = cashflow_hist.set_index("Year").to_dict(orient="index")

    g = assumptions["Revenue_Growth_YoY"]
    cogs_pct = assumptions["COGS_pct_of_Revenue"]
    opex_pct = assumptions["OpEx_pct_of_Revenue"]
    da_pct = assumptions["DA_pct_of_Revenue"]
    tax_rate = assumptions["Tax_Rate"]
    ar_pct = assumptions["AR_pct_of_Revenue"]
    inv_pct = assumptions["Inventory_pct_of_Revenue"]
    ap_pct = assumptions["AP_pct_of_Revenue"]
    capex_pct = assumptions["CapEx_pct_of_Revenue"]
    interest_rate = assumptions["Interest_Rate_on_Debt"]
    iss_pct = assumptions["Debt_Issuance_pct_of_Revenue"]
    rep_pct = assumptions["Debt_Repayment_pct_of_Revenue"]
    payout_pct = assumptions["Dividend_Payout_pct_of_NI"]

    last_year = max(hist_years)
    base_rev = inc_map[last_year]["Revenue"]
    cash0 = bs_map[last_year]["Cash"]
    debt0 = bs_map[last_year]["Debt"]
    equity_paid_in = bs_map[last_year]["Equity"]
    re0 = bs_map[last_year]["Retained_Earnings"]
    ar0 = bs_map[last_year]["Accounts_Receivable"]
    inv0 = bs_map[last_year]["Inventory"]
    ap0 = bs_map[last_year]["Accounts_Payable"]
    ppne0 = bs_map[last_year]["PP_and_E"]

    inc_rows, bs_rows, cf_rows, debt_sched = [], [], [], []
    prev_cash, prev_debt, prev_re = cash0, debt0, re0
    prev_ar, prev_inv, prev_ap = ar0, inv0, ap0
    prev_ppne = ppne0
    prev_rev = base_rev

    for i in range(n_years):
        y = start_year + i
        # Revenue
        rev = prev_rev * (1.0 + g)

        # Debt schedule (issuance/repay as % of rev)
        issuance = max(0.0, rev * iss_pct)
        repayment = max(0.0, rev * rep_pct)
        opening_debt = prev_debt
        ending_debt = opening_debt + issuance - repayment
        avg_debt = 0.5 * (opening_debt + ending_debt)
        interest_exp = avg_debt * interest_rate

        # IS
        cogs = rev * cogs_pct
        gross = rev - cogs
        opex = rev * opex_pct
        ebitda = gross - opex
        d_and_a = rev * da_pct
        ebit = ebitda - d_and_a
        ebt = ebit - interest_exp
        taxes = max(0.0, ebt * tax_rate)
        ni = ebt - taxes

        # WC balances
        ar = rev * ar_pct
        inv = rev * inv_pct
        ap = rev * ap_pct

        d_ar = -(ar - prev_ar)
        d_inv = -(inv - prev_inv)
        d_ap = (ap - prev_ap)

        cfo = ni + d_and_a + d_ar + d_inv + d_ap

        capex = rev * capex_pct
        ppne = prev_ppne + capex - d_and_a
        cfi = -capex

        dividends = -max(0.0, payout_pct * ni)
        cff = issuance - repayment + dividends  # (stock flows 0 by default)

        net_change = cfo + cfi + cff
        ending_cash = prev_cash + net_change

        re = prev_re + ni + dividends
        equity = equity_paid_in  # keep flat

        bs_rows.append({
            "Year": y,
            "Cash": ending_cash,
            "Accounts_Receivable": ar,
            "Inventory": inv,
            "PP_and_E": ppne,
            "Intangibles": 0.0,
            "Accounts_Payable": ap,
            "Debt": ending_debt,
            "Equity": equity,
            "Retained_Earnings": re,
            "Total_Assets": ending_cash + ar + inv + ppne,
            "Total_Liabilities": ap + ending_debt,
            "Total_Equity": equity + re,
            "Total_Liabilities_Equity": ap + ending_debt + equity + re
        })

        cf_rows.append({
            "Year": y,
            "Net_Income": ni,
            "Depreciation": d_and_a,
            "Change_in_AR": d_ar,
            "Change_in_Inventory": d_inv,
            "Change_in_AP": d_ap,
            "Cash_from_Operations": cfo,
            "CapEx": -capex,
            "Cash_from_Investing": -capex,
            "Debt_Issuance": issuance,
            "Debt_Repayment": -repayment,
            "Stock_Issuance": 0.0,
            "Stock_Buyback": 0.0,
            "Dividends": dividends,
            "Cash_from_Financing": issuance - repayment + dividends,
            "Net_Change_in_Cash": net_change,
            "Beginning_Cash": prev_cash,
            "Ending_Cash": ending_cash
        })

        inc_rows.append({
            "Year": y,
            "Revenue": rev,
            "COGS": cogs,
            "Gross_Profit": gross,
            "Operating_Expenses": opex,
            "EBITDA": ebitda,
            "Depreciation": d_and_a,
            "EBIT": ebit,
            "Interest_Expense": interest_exp,
            "EBT": ebt,
            "Taxes": taxes,
            "Net_Income": ni,
            "Tax_Rate": tax_rate
        })

        debt_sched.append({
            "Year": y,
            "Opening_Debt": opening_debt,
            "Issuance": issuance,
            "Repayment": -repayment,
            "Ending_Debt": ending_debt,
            "Avg_Debt": avg_debt,
            "Interest_Rate_Assumed": interest_rate,
            "Interest_Expense": interest_exp
        })

        # roll
        prev_cash, prev_debt, prev_re = ending_cash, ending_debt, re
        prev_ar, prev_inv, prev_ap = ar, inv, ap
        prev_ppne = ppne
        prev_rev = rev

    return (pd.DataFrame(inc_rows), pd.DataFrame(bs_rows),
            pd.DataFrame(cf_rows), pd.DataFrame(debt_sched))

# ==========================
# Public loader (2020–2029) + assumptions reuse
# ==========================
def load_financial_data(ticker, reuse_assumptions=False, assumptions_path=None):
    cik = get_cik(ticker)
    facts = _get_json(_company_facts_url(cik))
    time.sleep(RATE_SLEEP)

    actual_years = [2020, 2021, 2022, 2023, 2024]
    income_hist = _build_historical_income(facts, actual_years)
    balance_hist, cashflow_hist = _build_historical_bs_cf(facts, actual_years, income_hist)

    # Assumptions (defaults inferred)
    assumptions = infer_default_assumptions(income_hist, balance_hist, cashflow_hist)

    # If reuse requested and an existing Excel with Assumptions exists, read it and override
    if reuse_assumptions:
        assumed = try_read_assumptions_from_excel(assumptions_path or f"{ticker}_three_statement_model_2020_2029.xlsx")
        if assumed:
            assumptions.update(assumed)

    # Projections with assumptions (includes debt schedule)
    inc_proj, bs_proj, cf_proj, debt_sched = project_forward_with_assumptions(
        income_hist, balance_hist, cashflow_hist, assumptions, start_year=2025, n_years=5
    )

    income_df = pd.concat([income_hist, inc_proj], ignore_index=True)
    balance_df = pd.concat([balance_hist, bs_proj], ignore_index=True)
    cashflow_df = pd.concat([cashflow_hist, cf_proj], ignore_index=True)

    return income_df, balance_df, cashflow_df, assumptions, debt_sched

def try_read_assumptions_from_excel(xlsx_path):
    try:
        if not os.path.exists(xlsx_path):
            return None
        wb = load_workbook(xlsx_path, data_only=True)
        if "Assumptions" not in wb.sheetnames:
            return None
        ws = wb["Assumptions"]
        out = {}
        for r in range(2, ws.max_row + 1):
            key = ws.cell(row=r, column=1).value
            val = ws.cell(row=r, column=2).value
            if key is None:
                continue
            try:
                # accept numeric
                v = float(val)
            except (TypeError, ValueError):
                v = val
            out[str(key)] = v
        return out
    except Exception:
        return None

# ==========================
# Validation & integrated model (unchanged)
# ==========================
def validate_input_data(income_df, balance_df, cashflow_df):
    req_is = ['Year', 'Revenue', 'COGS', 'Operating_Expenses', 'Depreciation', 'Interest_Expense', 'Tax_Rate']
    req_bs = ['Year', 'Cash', 'Accounts_Receivable', 'Inventory', 'PP_and_E', 'Intangibles',
              'Accounts_Payable', 'Debt', 'Equity', 'Retained_Earnings']
    req_cf = ['Year', 'CapEx', 'Dividends', 'Debt_Issuance', 'Debt_Repayment', 'Stock_Issuance', 'Stock_Buyback']
    for df, req in [(income_df, req_is), (balance_df, req_bs), (cashflow_df, req_cf)]:
        for c in req:
            if c not in df.columns:
                df[c] = 0
        df["Year"] = df["Year"].astype(int)
    yrs = set(income_df["Year"]) & set(balance_df["Year"]) & set(cashflow_df["Year"])
    if len(yrs) < 2:
        raise ValueError(f"Insufficient overlapping years: {sorted(yrs)}")
    keep = sorted(yrs)
    return (income_df[income_df["Year"].isin(keep)].copy(),
            balance_df[balance_df["Year"].isin(keep)].copy(),
            cashflow_df[cashflow_df["Year"].isin(keep)].copy())

def create_three_statement_model(income_df, balance_df, cashflow_df):
    years = sorted(income_df['Year'].unique())
    income_model, balance_model, cashflow_model = [], [], []

    for i, year in enumerate(years):
        inc = income_df[income_df['Year'] == year].iloc[0]
        bal_prev = balance_df[balance_df['Year'] == years[i-1]].iloc[0] if i > 0 else None
        bal_cur_in = balance_df[balance_df['Year'] == year].iloc[0]
        cf_in = cashflow_df[cashflow_df['Year'] == year].iloc[0]

        row_i = {
            'Year': int(year),
            'Revenue': inc.get('Revenue', 0),
            'COGS': inc.get('COGS', 0),
            'Gross_Profit': inc.get('Revenue', 0) - inc.get('COGS', 0),
            'Operating_Expenses': inc.get('Operating_Expenses', 0),
            'EBITDA': (inc.get('Revenue', 0) - inc.get('COGS', 0)) - inc.get('Operating_Expenses', 0),
            'Depreciation': inc.get('Depreciation', 0),
            'EBIT': ((inc.get('Revenue', 0) - inc.get('COGS', 0)) - inc.get('Operating_Expenses', 0)) - inc.get('Depreciation', 0),
            'Interest_Expense': inc.get('Interest_Expense', 0),
        }
        row_i['EBT'] = row_i['EBIT'] - row_i['Interest_Expense']
        tax_rate = inc.get('Tax_Rate', 0.21)
        row_i['Taxes'] = row_i['EBT'] * tax_rate
        row_i['Net_Income'] = row_i['EBT'] - row_i['Taxes']
        income_model.append(row_i)

        row_c = {
            'Year': int(year),
            'Net_Income': row_i['Net_Income'],
            'Depreciation': row_i['Depreciation'],
        }
        prev_ar = bal_prev.get('Accounts_Receivable', 0) if bal_prev is not None else 0
        prev_inv = bal_prev.get('Inventory', 0) if bal_prev is not None else 0
        prev_ap = bal_prev.get('Accounts_Payable', 0) if bal_prev is not None else 0
        cur_ar = bal_cur_in.get('Accounts_Receivable', 0)
        cur_inv = bal_cur_in.get('Inventory', 0)
        cur_ap = bal_cur_in.get('Accounts_Payable', 0)

        row_c['Change_in_AR'] = -(cur_ar - prev_ar)
        row_c['Change_in_Inventory'] = -(cur_inv - prev_inv)
        row_c['Change_in_AP'] = (cur_ap - prev_ap)

        row_c['Cash_from_Operations'] = (
            row_c['Net_Income'] + row_c['Depreciation'] +
            row_c['Change_in_AR'] + row_c['Change_in_Inventory'] + row_c['Change_in_AP']
        )

        row_c['CapEx'] = cf_in.get('CapEx', 0) if cf_in.get('CapEx', 0) <= 0 else -abs(cf_in.get('CapEx', 0))
        row_c['Cash_from_Investing'] = row_c['CapEx']
        row_c['Debt_Issuance'] = cf_in.get('Debt_Issuance', 0)
        row_c['Debt_Repayment'] = cf_in.get('Debt_Repayment', 0) if cf_in.get('Debt_Repayment', 0) <= 0 else -abs(cf_in.get('Debt_Repayment', 0))
        row_c['Stock_Issuance'] = cf_in.get('Stock_Issuance', 0)
        row_c['Stock_Buyback'] = cf_in.get('Stock_Buyback', 0) if cf_in.get('Stock_Buyback', 0) <= 0 else -abs(cf_in.get('Stock_Buyback', 0))
        row_c['Dividends'] = cf_in.get('Dividends', 0) if cf_in.get('Dividends', 0) <= 0 else -abs(cf_in.get('Dividends', 0))

        row_c['Cash_from_Financing'] = (
            row_c['Debt_Issuance'] + row_c['Debt_Repayment'] +
            row_c['Stock_Issuance'] + row_c['Stock_Buyback'] + row_c['Dividends']
        )
        row_c['Net_Change_in_Cash'] = row_c['Cash_from_Operations'] + row_c['Cash_from_Investing'] + row_c['Cash_from_Financing']
        beginning_cash = bal_prev.get('Cash', 0) if bal_prev is not None else bal_cur_in.get('Cash', 0)
        row_c['Beginning_Cash'] = beginning_cash
        row_c['Ending_Cash'] = beginning_cash + row_c['Net_Change_in_Cash']
        cashflow_model.append(row_c)

        row_b = {
            'Year': int(year),
            'Cash': row_c['Ending_Cash'],
            'Accounts_Receivable': cur_ar,
            'Inventory': cur_inv,
            'PP_and_E': bal_cur_in.get('PP_and_E', 0),
            'Intangibles': bal_cur_in.get('Intangibles', 0),
        }
        row_b['Total_Assets'] = row_b['Cash'] + row_b['Accounts_Receivable'] + row_b['Inventory'] + row_b['PP_and_E'] + row_b['Intangibles']
        row_b['Accounts_Payable'] = cur_ap
        row_b['Debt'] = bal_cur_in.get('Debt', 0)
        row_b['Total_Liabilities'] = row_b['Accounts_Payable'] + row_b['Debt']
        row_b['Equity'] = bal_cur_in.get('Equity', 0)
        prev_re = bal_prev.get('Retained_Earnings', 0) if bal_prev is not None else 0
        dividends_abs = abs(row_c['Dividends'])
        row_b['Retained_Earnings'] = prev_re + row_i['Net_Income'] - dividends_abs
        row_b['Total_Equity'] = row_b['Equity'] + row_b['Retained_Earnings']
        row_b['Total_Liabilities_Equity'] = row_b['Total_Liabilities'] + row_b['Total_Equity']
        balance_model.append(row_b)

        assets = row_b['Total_Assets']; liab_eq = row_b['Total_Liabilities_Equity']
        if not np.isclose(assets, liab_eq, rtol=0.05):
            print(f"Warning: Balance Sheet imbalance for {year}: Assets {assets:,.0f} vs L+E {liab_eq:,.0f}")

    return pd.DataFrame(income_model), pd.DataFrame(balance_model), pd.DataFrame(cashflow_model)

# ==========================
# Excel writer (labels in col A + Assumptions + Debt Schedule)
# ==========================
def save_formatted_excel(income_df, balance_df, cashflow_df, assumptions, debt_sched, filename, ticker, company_title=None):
    IS_ORDER = [
        ("Revenue", "Revenue"), ("COGS", "COGS"), ("Gross_Profit", "Gross Profit"),
        ("Operating_Expenses", "Operating Expenses"), ("EBITDA", "EBITDA"),
        ("Depreciation", "Depreciation & Amortization"), ("EBIT", "EBIT"),
        ("Interest_Expense", "Interest Expense"), ("EBT", "EBT"),
        ("Taxes", "Taxes"), ("Net_Income", "Net Income"),
    ]
    BS_ORDER = [
        ("Cash", "Cash & Cash Equivalents"), ("Accounts_Receivable", "Accounts Receivable"),
        ("Inventory", "Inventory"), ("PP_and_E", "Property, Plant & Equipment"), ("Intangibles", "Intangibles"),
        ("Total_Assets", "Total Assets"), ("Accounts_Payable", "Accounts Payable"), ("Debt", "Total Debt"),
        ("Total_Liabilities", "Total Liabilities"), ("Equity", "Common Equity (Paid-in Capital)"),
        ("Retained_Earnings", "Retained Earnings"), ("Total_Equity", "Total Equity"),
        ("Total_Liabilities_Equity", "Total Liabilities & Equity"),
    ]
    CF_ORDER = [
        ("Net_Income", "Net Income"), ("Depreciation", "Depreciation & Amortization"),
        ("Change_in_AR", "Change in Accounts Receivable"), ("Change_in_Inventory", "Change in Inventory"),
        ("Change_in_AP", "Change in Accounts Payable"), ("Cash_from_Operations", "Cash from Operating Activities"),
        ("CapEx", "Capital Expenditures (CapEx)"), ("Cash_from_Investing", "Cash from Investing Activities"),
        ("Debt_Issuance", "Debt Issuance"), ("Debt_Repayment", "Debt Repayment"),
        ("Stock_Issuance", "Stock Issuance"), ("Stock_Buyback", "Stock Buybacks"),
        ("Dividends", "Dividends"), ("Cash_from_Financing", "Cash from Financing Activities"),
        ("Net_Change_in_Cash", "Net Change in Cash"), ("Beginning_Cash", "Beginning Cash"),
        ("Ending_Cash", "Ending Cash"),
    ]

    def _years(df): return sorted(df["Year"].astype(int).unique())
    def _lookup(df):
        return {int(r["Year"]): r.to_dict() for _, r in df.iterrows()}

    wb = Workbook()
    wb.remove(wb.active)

    ws_income = wb.create_sheet(f"{ticker} - Income Statement")
    ws_balance = wb.create_sheet(f"{ticker} - Balance Sheet")
    ws_cash   = wb.create_sheet(f"{ticker} - Cash Flow Statement")
    ws_assump = wb.create_sheet("Assumptions")
    ws_debt   = wb.create_sheet("Debt Schedule")

    title_font = Font(bold=True, size=16, color="000080")
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    cell_align  = Alignment(horizontal="center", vertical="center")
    left_align  = Alignment(horizontal="left", vertical="center")
    border = Border(left=Side(style="thin"), right=Side(style="thin"),
                    top=Side(style="thin"), bottom=Side(style="thin"))

    def write_statement(ws, sheet_title, order, df):
        ws.cell(row=1, column=1, value=sheet_title if not company_title else f"{sheet_title} — {company_title}").font = title_font
        ws.cell(row=1, column=1).alignment = Alignment(horizontal="center", vertical="center")

        ws.cell(row=2, column=1, value="Line Item")
        ws.cell(row=2, column=1).font = header_font
        ws.cell(row=2, column=1).fill = header_fill
        ws.cell(row=2, column=1).alignment = left_align
        ws.cell(row=2, column=1).border = border

        years = _years(df)
        for j, yr in enumerate(years, start=2):
            c = ws.cell(row=2, column=j, value=int(yr))
            c.font = header_font; c.fill = header_fill; c.alignment = cell_align; c.border = border

        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1 + len(years))

        m = _lookup(df)
        r = 3
        for key, label in order:
            ws.cell(row=r, column=1, value=label).alignment = left_align
            ws.cell(row=r, column=1).border = border
            for j, yr in enumerate(years, start=2):
                v = m.get(yr, {}).get(key, None)
                cell = ws.cell(row=r, column=j, value=v)
                cell.alignment = cell_align; cell.border = border
                if isinstance(v, (int, float)): cell.number_format = '$#,##0;($#,##0)'
            r += 1

        ws.column_dimensions[get_column_letter(1)].width = 40
        for col in range(2, 2 + len(years)):
            ws.column_dimensions[get_column_letter(col)].width = 18

    # Statements
    write_statement(ws_income, "Income Statement", IS_ORDER, income_df)
    write_statement(ws_balance, "Balance Sheet", BS_ORDER, balance_df)
    write_statement(ws_cash,   "Cash Flow Statement", CF_ORDER, cashflow_df)

    # Assumptions sheet (KVP)
    ws_assump.cell(row=1, column=1, value="Assumption").font = header_font
    ws_assump.cell(row=1, column=2, value="Value").font = header_font
    ws_assump.cell(row=1, column=1).fill = header_fill
    ws_assump.cell(row=1, column=2).fill = header_fill
    ws_assump.column_dimensions["A"].width = 40
    ws_assump.column_dimensions["B"].width = 18
    r = 2
    for k, v in assumptions.items():
        ws_assump.cell(row=r, column=1, value=k)
        ws_assump.cell(row=r, column=2, value=float(v) if isinstance(v, (int, float)) else v)
        if isinstance(v, (int, float)):
            ws_assump.cell(row=r, column=2).number_format = '0.0000'
        r += 1

    # Debt Schedule sheet
    ws_debt.cell(row=1, column=1, value="Debt Schedule").font = title_font
    cols = ["Year", "Opening_Debt", "Issuance", "Repayment", "Ending_Debt", "Avg_Debt", "Interest_Rate_Assumed", "Interest_Expense"]
    for j, name in enumerate(cols, start=1):
        c = ws_debt.cell(row=2, column=j, value=name); c.font = header_font; c.fill = header_fill
    row = 3
    for _, rec in debt_sched.iterrows():
        for j, name in enumerate(cols, start=1):
            val = rec[name]
            cell = ws_debt.cell(row=row, column=j, value=val)
            if name != "Year":
                cell.number_format = '$#,##0;($#,##0)' if "Rate" not in name else '0.0000'
        row += 1
    for col in range(1, len(cols)+1):
        ws_debt.column_dimensions[get_column_letter(col)].width = 18

    wb.save(filename)

# ==========================
# PDF export (summary)
# ==========================
def export_summary_pdf(pdf_path, ticker, company_title, income_model, balance_model, cashflow_model, assumptions):
    """
    Create an analysis-focused PDF:
      - Narrative analysis sections with computed metrics
      - KPI snapshot + Three-statement overview (latest year)
      - Cash Flow Waterfall (latest year)
      - Assumptions table

    If 'reportlab' isn't installed, the function returns False and prints a tip.
    """
    try:
        from reportlab.lib.pagesizes import LETTER
        from reportlab.pdfgen import canvas
        from reportlab.lib.units import inch
        from reportlab.lib import colors

        # ---------- helpers ----------
        def fmt_currency(v):
            try: return f"${float(v):,.0f}"
            except: return "-"

        def fmt_pct(v, digits=1):
            try: return f"{float(v)*100:,.{digits}f}%"
            except: return "-"

        def fmt_any(v):
            return fmt_currency(v) if isinstance(v, (int, float)) else str(v)

        def safe_float(v, default=0.0):
            try:
                f = float(v)
                if f == float("inf") or f != f:  # inf or NaN
                    return default
                return f
            except Exception:
                return default

        # text layout helpers
        PAGE_W, PAGE_H = LETTER
        LEFT = 0.75 * inch
        RIGHT = PAGE_W - 0.75 * inch
        TOP = PAGE_H - 0.75 * inch
        BOTTOM = 0.75 * inch
        LINE_GAP = 14
        SEC_GAP = 18

        c = canvas.Canvas(pdf_path, pagesize=LETTER)
        page_num = 1

        def footer():
            c.setFont("Helvetica", 8)
            c.setFillColor(colors.grey)
            c.drawRightString(RIGHT, 0.5 * inch, f"Page {page_num}")
            c.setFillColor(colors.black)

        def hrule(y):
            c.setStrokeColor(colors.lightgrey)
            c.setLineWidth(0.6)
            c.line(LEFT, y, RIGHT, y)
            c.setStrokeColor(colors.black)

        def section_title(text, y):
            c.setFont("Helvetica-Bold", 12)
            c.drawString(LEFT, y, text)
            return y - LINE_GAP

        def ensure_space(y, needed=100):
            nonlocal page_num
            if y - needed < BOTTOM:
                footer()
                c.showPage()
                page_num += 1
                return TOP
            return y

        def wrap_paragraph(text, x, y, max_width, line_gap=LINE_GAP, font="Helvetica", size=10):
            """Simple word-wrap onto the canvas; returns new y."""
            c.setFont(font, size)
            words = str(text).split()
            line = ""
            for w in words:
                trial = (line + " " + w).strip()
                if c.stringWidth(trial, font, size) <= max_width:
                    line = trial
                else:
                    c.drawString(x, y, line)
                    y -= line_gap
                    if y < BOTTOM + line_gap:
                        # new page if needed
                        y = ensure_space(y, needed=TOP - BOTTOM)
                    line = w
            if line:
                c.drawString(x, y, line)
                y -= line_gap
            return y

        def draw_bullets(items, x, y, max_width, bullet="•", indent=10, line_gap=LINE_GAP):
            for it in items:
                y = ensure_space(y, needed=2*line_gap)
                c.drawString(x, y, bullet)
                y = wrap_paragraph(it, x + indent, y, max_width - indent, line_gap=line_gap)
            return y

        def draw_kv_table(rows, y_start, col_split=3.8*inch, label_width=None):
            y = y_start
            c.setFont("Helvetica", 10)
            for label, value in rows:
                y = ensure_space(y, needed=LINE_GAP)
                lbl, val = str(label), str(value)
                c.drawString(LEFT, y, lbl)
                # dotted leaders
                text_w = c.stringWidth(lbl, "Helvetica", 10)
                max_lbl_w = (label_width if label_width is not None else (col_split - LEFT - 8))
                dots_start = LEFT + min(text_w + 6, max_lbl_w)
                c.setStrokeColor(colors.lightgrey)
                c.setDash(1, 2); c.line(dots_start, y - 2, col_split - 6, y - 2); c.setDash()
                c.setStrokeColor(colors.black)
                # value right-aligned
                c.drawRightString(RIGHT, y, val)
                y -= LINE_GAP
            return y

        # ---------- gather historical + latest ----------
        years = sorted(set(int(y) for y in income_model["Year"]))
        hist_years = [y for y in years if y <= 2024]
        last_y = max(years)
        first_hist = min(hist_years) if hist_years else years[0]
        im_last = income_model[income_model["Year"] == last_y].iloc[0]
        bm_last = balance_model[balance_model["Year"] == last_y].iloc[0]
        cf_last = cashflow_model[cashflow_model["Year"] == last_y].iloc[0]

        # Historical endpoints for trends
        im_first = income_model[income_model["Year"] == first_hist].iloc[0]
        bm_first = balance_model[balance_model["Year"] == first_hist].iloc[0]

        # ---------- compute analysis metrics ----------
        rev_0 = safe_float(im_first.get("Revenue"))
        rev_T = safe_float(im_last.get("Revenue"))
        periods = max(1, (2024 - first_hist) if last_y >= 2025 else (last_y - first_hist))
        rev_cagr = (rev_T / rev_0) ** (1/periods) - 1 if rev_0 > 0 and rev_T > 0 else 0.0

        gpm_0  = (safe_float(im_first.get("Gross_Profit")) / rev_0) if rev_0 else 0.0
        gpm_T  = (safe_float(im_last.get("Gross_Profit")) / rev_T) if rev_T else 0.0
        ebit_T = safe_float(im_last.get("EBIT"))
        ebit_m_T = (ebit_T / rev_T) if rev_T else 0.0
        ni_m_T = (safe_float(im_last.get("Net_Income")) / rev_T) if rev_T else 0.0

        # Cash flow quality
        cfo_T = safe_float(cf_last.get("Cash_from_Operations"))
        ni_T  = safe_float(im_last.get("Net_Income"))
        cfo_ni_ratio = (cfo_T / ni_T) if ni_T != 0 else 0.0
        capex_T = -safe_float(cf_last.get("CapEx")) if safe_float(cf_last.get("CapEx")) < 0 else safe_float(cf_last.get("CapEx"))  # positive number
        fcf_T = cfo_T - capex_T
        fcf_margin_T = (fcf_T / rev_T) if rev_T else 0.0

        # Investment intensity
        capex_pct_hist = []
        for y in hist_years:
            rv = safe_float(income_model.loc[income_model["Year"] == y, "Revenue"].values[0])
            cap = -safe_float(cashflow_model.loc[cashflow_model["Year"] == y, "CapEx"].values[0])
            if rv > 0:
                capex_pct_hist.append(cap/rv)
        capex_avg_hist = sum(capex_pct_hist)/len(capex_pct_hist) if capex_pct_hist else 0.0

        # Capital structure & coverage
        debt_T = safe_float(bm_last.get("Debt"))
        cash_T = safe_float(bm_last.get("Cash"))
        net_debt_T = debt_T - cash_T
        ebitda_T = safe_float(im_last.get("EBITDA"))
        debt_ebitda_T = (debt_T / ebitda_T) if ebitda_T > 0 else None
        int_exp_T = safe_float(im_last.get("Interest_Expense"))
        int_cov_T = (ebit_T / int_exp_T) if int_exp_T > 0 else None

        # Working capital efficiency (days)
        cogs_T = safe_float(im_last.get("COGS"))
        ar_T = safe_float(bm_last.get("Accounts_Receivable"))
        inv_T = safe_float(bm_last.get("Inventory"))
        ap_T = safe_float(bm_last.get("Accounts_Payable"))
        DSO = (ar_T / rev_T) * 365 if rev_T > 0 else None
        DIO = (inv_T / cogs_T) * 365 if cogs_T > 0 else None
        DPO = (ap_T / cogs_T) * 365 if cogs_T > 0 else None
        CCC = (DSO or 0) + (DIO or 0) - (DPO or 0) if None not in (DSO, DIO, DPO) else None

        # Forward outlook from assumptions
        g_ass    = float(assumptions.get("Revenue_Growth_YoY", 0.0) or 0.0)
        cogs_pct = float(assumptions.get("COGS_pct_of_Revenue", 0.0) or 0.0)
        opex_pct = float(assumptions.get("OpEx_pct_of_Revenue", 0.0) or 0.0)
        da_pct   = float(assumptions.get("DA_pct_of_Revenue", 0.0) or 0.0)
        capex_pct= float(assumptions.get("CapEx_pct_of_Revenue", 0.0) or 0.0)
        tax_rate = float(assumptions.get("Tax_Rate", 0.0) or 0.0)
        int_pct  = float(assumptions.get("Interest_Rate_on_Debt", 0.04) or 0.04)  # if used as % of avg debt; for a margin proxy we’ll approximate % of rev below

        ebitda_margin_proxy = 1 - cogs_pct - opex_pct
        ebit_margin_proxy   = ebitda_margin_proxy - da_pct
        # NI margin proxy: EBIT – interest (approx via interest as % of rev if you set Debt Iss/Rep to 0) – taxes
        # We’ll use a conservative interest impact of 0.5% of revenue if not specified separately:
        interest_as_rev_pct = float(assumptions.get("Interest_pct_of_Revenue", 0.005) or 0.005)
        ebt_margin_proxy = ebit_margin_proxy - interest_as_rev_pct
        ni_margin_proxy  = ebt_margin_proxy * (1 - tax_rate)
        fcf_margin_proxy = ni_margin_proxy + da_pct - capex_pct  # simple FCF proxy

        # ---------- Header ----------
        y = TOP
        c.setFont("Helvetica-Bold", 16)
        title = f"{ticker} — Three-Statement Model (2020–2029)"
        c.drawString(LEFT, y, title)
        if company_title:
            c.setFont("Helvetica", 10)
            c.setFillColor(colors.darkgrey)
            c.drawString(LEFT, y - 16, f"Company: {company_title}")
            c.setFillColor(colors.black)
            y -= 8
        y -= SEC_GAP
        hrule(y); y -= SEC_GAP

        # ---------- Overview (narrative) ----------
        y = section_title("Overview", y)
        overview_txt = (
            f"Over the analyzed period, revenue grew at a CAGR of {fmt_pct(rev_cagr)} from "
            f"{fmt_currency(rev_0)} to {fmt_currency(rev_T)}. Latest-year margins show "
            f"gross margin at {fmt_pct(gpm_T)}, EBIT margin at {fmt_pct(ebit_m_T)}, and "
            f"net income margin at {fmt_pct(ni_m_T)}."
        )
        y = wrap_paragraph(overview_txt, LEFT, y, RIGHT-LEFT)

        y = section_title("Performance Highlights", y)
        bullets = [
            f"Gross margin {'expanded' if gpm_T >= gpm_0 else 'compressed'} by "
            f"{fmt_pct(abs(gpm_T - gpm_0))} vs {first_hist}.",
            f"EBIT margin (latest): {fmt_pct(ebit_m_T)}; Net margin: {fmt_pct(ni_m_T)}.",
            f"Free cash flow margin (latest): {fmt_pct(fcf_margin_T)}; "
            f"Cash-from-ops to net income ratio at {fmt_pct(cfo_ni_ratio, digits=0)} "
            f"({'>=100%' if cfo_ni_ratio >= 1 else '<100%'} signals "
            f"{'strong' if cfo_ni_ratio >= 1 else 'weaker'} cash conversion).",
        ]
        y = draw_bullets(bullets, LEFT, y, RIGHT-LEFT)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Investment & CapEx ----------
        y = section_title("Investment & CapEx", y)
        invest_txt = (
            f"CapEx averaged {fmt_pct(capex_avg_hist)} of revenue across historical years; "
            f"in the latest year it was {fmt_pct(capex_T / rev_T if rev_T else 0)}. "
            f"This implies {'steady' if abs((capex_T/rev_T if rev_T else 0) - capex_avg_hist) < 0.01 else 'changing'} reinvestment intensity. "
            "PP&E trends, combined with D&A, suggest the asset base is "
            f"{'expanding' if safe_float(bm_last.get('PP_and_E')) > safe_float(bm_first.get('PP_and_E')) else 'stable/shrinking'}."
        )
        y = wrap_paragraph(invest_txt, LEFT, y, RIGHT-LEFT)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Capital structure ----------
        y = section_title("Capital Structure & Coverage", y)
        cap_bullets = [
            f"Net debt (latest): {fmt_currency(net_debt_T)} (Debt {fmt_currency(debt_T)} − Cash {fmt_currency(cash_T)}).",
            f"Debt / EBITDA: {('N/A' if debt_ebitda_T is None else f'{debt_ebitda_T:,.2f}×')}, "
            f"Interest coverage (EBIT / Interest): {('N/A' if int_cov_T is None else f'{int_cov_T:,.1f}×')}."
        ]
        y = draw_bullets(cap_bullets, LEFT, y, RIGHT-LEFT)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Working capital ----------
        y = section_title("Working Capital Efficiency", y)
        wc_bullets = [
            f"DSO: {('N/A' if DSO is None else f'{DSO:,.0f} days')}, "
            f"DIO: {('N/A' if DIO is None else f'{DIO:,.0f} days')}, "
            f"DPO: {('N/A' if DPO is None else f'{DPO:,.0f} days')}.",
            f"Cash conversion cycle: {('N/A' if CCC is None else f'{CCC:,.0f} days')} "
            "(lower is better; negative CCC indicates supplier financing advantage)."
        ]
        y = draw_bullets(wc_bullets, LEFT, y, RIGHT-LEFT)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Forward outlook ----------
        y = section_title("Forward Outlook (based on current assumptions)", y)
        outlook_txt = (
            f"Model assumes revenue grows {fmt_pct(g_ass)} annually. "
            f"Projected EBITDA margin: {fmt_pct(ebitda_margin_proxy)}, "
            f"EBIT margin: {fmt_pct(ebit_margin_proxy)}, "
            f"and a simplified free cash flow margin near {fmt_pct(fcf_margin_proxy)} "
            "(= NI margin + D&A% − CapEx%). Adjust growth, margin, CapEx, and WC "
            "percentages in the Assumptions tab to see how this shifts cash generation and leverage."
        )
        y = wrap_paragraph(outlook_txt, LEFT, y, RIGHT-LEFT)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- KPI Snapshot (latest) ----------
        y = section_title(f"KPI Snapshot — {last_y}", y)
        kpi_rows = [
            ("Revenue", fmt_currency(im_last.get("Revenue", 0))),
            ("Gross Profit", fmt_currency(im_last.get("Gross_Profit", 0))),
            ("EBITDA", fmt_currency(im_last.get("EBITDA", 0))),
            ("EBIT", fmt_currency(im_last.get("EBIT", 0))),
            ("Net Income", fmt_currency(im_last.get("Net_Income", 0))),
            ("Cash (End)", fmt_currency(cf_last.get("Ending_Cash", bm_last.get("Cash", 0)))),
            ("Total Assets", fmt_currency(bm_last.get("Total_Assets", 0))),
            ("Liabilities + Equity", fmt_currency(bm_last.get("Total_Liabilities_Equity", 0))),
        ]
        y = draw_kv_table(kpi_rows, y)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Three-Statement Overview (latest) ----------
        y = section_title("Three-Statement Overview (latest year)", y)

        # Income Statement
        c.setFont("Helvetica-Bold", 11); c.drawString(LEFT, y, "Income Statement"); y -= LINE_GAP
        is_rows = [
            ("Revenue", fmt_currency(im_last.get("Revenue", 0))),
            ("COGS", fmt_currency(im_last.get("COGS", 0))),
            ("Gross Profit", fmt_currency(im_last.get("Gross_Profit", 0))),
            ("Operating Expenses", fmt_currency(im_last.get("Operating_Expenses", 0))),
            ("EBITDA", fmt_currency(im_last.get("EBITDA", 0))),
            ("Depreciation & Amortization", fmt_currency(im_last.get("Depreciation", 0))),
            ("EBIT", fmt_currency(im_last.get("EBIT", 0))),
            ("Interest Expense", fmt_currency(im_last.get("Interest_Expense", 0))),
            ("EBT", fmt_currency(im_last.get("EBT", 0))),
            ("Taxes", fmt_currency(im_last.get("Taxes", 0))),
            ("Net Income", fmt_currency(im_last.get("Net_Income", 0))),
        ]
        y = draw_kv_table(is_rows, y); y -= 4

        # Balance Sheet
        y = ensure_space(y, needed=220)
        c.setFont("Helvetica-Bold", 11); c.drawString(LEFT, y, "Balance Sheet"); y -= LINE_GAP
        bs_rows = [
            ("Cash & Cash Equivalents", fmt_currency(bm_last.get("Cash", 0))),
            ("Accounts Receivable", fmt_currency(bm_last.get("Accounts_Receivable", 0))),
            ("Inventory", fmt_currency(bm_last.get("Inventory", 0))),
            ("PP&E (Net)", fmt_currency(bm_last.get("PP_and_E", 0))),
            ("Intangibles", fmt_currency(bm_last.get("Intangibles", 0))),
            ("Total Assets", fmt_currency(bm_last.get("Total_Assets", 0))),
            ("Accounts Payable", fmt_currency(bm_last.get("Accounts_Payable", 0))),
            ("Total Debt", fmt_currency(bm_last.get("Debt", 0))),
            ("Total Liabilities", fmt_currency(bm_last.get("Total_Liabilities", 0))),
            ("Paid-in Capital", fmt_currency(bm_last.get("Equity", 0))),
            ("Retained Earnings", fmt_currency(bm_last.get("Retained_Earnings", 0))),
            ("Total Equity", fmt_currency(bm_last.get("Total_Equity", 0))),
            ("Liabilities + Equity", fmt_currency(bm_last.get("Total_Liabilities_Equity", 0))),
        ]
        y = draw_kv_table(bs_rows, y); y -= 4

        # Cash Flow
        y = ensure_space(y, needed=220)
        c.setFont("Helvetica-Bold", 11); c.drawString(LEFT, y, "Cash Flow"); y -= LINE_GAP
        cf_rows = [
            ("Net Income", fmt_currency(cf_last.get("Net_Income", 0))),
            ("+ Depreciation & Amortization", fmt_currency(cf_last.get("Depreciation", 0))),
            ("± Change in A/R", fmt_currency(cf_last.get("Change_in_AR", 0))),
            ("± Change in Inventory", fmt_currency(cf_last.get("Change_in_Inventory", 0))),
            ("± Change in A/P", fmt_currency(cf_last.get("Change_in_AP", 0))),
            ("Cash from Operations (CFO)", fmt_currency(cf_last.get("Cash_from_Operations", 0))),
            ("CapEx", fmt_currency(cf_last.get("CapEx", 0))),
            ("Cash from Investing (CFI)", fmt_currency(cf_last.get("Cash_from_Investing", 0))),
            ("Debt Issuance", fmt_currency(cf_last.get("Debt_Issuance", 0))),
            ("Debt Repayment", fmt_currency(cf_last.get("Debt_Repayment", 0))),
            ("Stock Issuance", fmt_currency(cf_last.get("Stock_Issuance", 0))),
            ("Stock Buybacks", fmt_currency(cf_last.get("Stock_Buyback", 0))),
            ("Dividends", fmt_currency(cf_last.get("Dividends", 0))),
            ("Cash from Financing (CFF)", fmt_currency(cf_last.get("Cash_from_Financing", 0))),
            ("Net Change in Cash", fmt_currency(cf_last.get("Net_Change_in_Cash", 0))),
            ("Beginning Cash", fmt_currency(cf_last.get("Beginning_Cash", 0))),
            ("Ending Cash", fmt_currency(cf_last.get("Ending_Cash", 0))),
        ]
        y = draw_kv_table(cf_rows, y)
        y -= 6
        hrule(y); y -= SEC_GAP

        # ---------- Cash Flow Waterfall (latest year) ----------
        y = section_title("Cash Flow Waterfall (latest year)", y)
        y = ensure_space(y, needed=220)

        beg_cash = safe_float(cf_last.get("Beginning_Cash"))
        cfo      = safe_float(cf_last.get("Cash_from_Operations"))
        cfi      = safe_float(cf_last.get("Cash_from_Investing"))
        cff      = safe_float(cf_last.get("Cash_from_Financing"))
        netchg   = safe_float(cf_last.get("Net_Change_in_Cash", cfo + cfi + cff))
        end_cash = safe_float(cf_last.get("Ending_Cash", beg_cash + netchg))

        bars = [
            ("Beginning Cash", beg_cash, "base"),
            ("CFO", cfo, "flow"),
            ("CFI", cfi, "flow"),
            ("CFF", cff, "flow"),
            ("Net Δ Cash", netchg, "sum"),
            ("Ending Cash", end_cash, "base"),
        ]

        chart_left  = LEFT
        chart_right = RIGHT
        chart_width = chart_right - chart_left
        chart_bottom = y - 160
        chart_top    = y - 20
        chart_height = chart_top - chart_bottom

        cum = beg_cash
        cum_values = [beg_cash]
        for _, v, kind in bars[1:]:
            if kind == "flow":
                cum += v
            elif kind == "sum":
                cum = beg_cash + (cfo + cfi + cff)
            elif kind == "base":
                cum = end_cash
            cum_values.append(cum)

        max_abs = max(1.0, max(abs(x) for x in [beg_cash, cfo, cfi, cff, netchg, end_cash] + cum_values))
        def y_scale(val):
            zero_y = chart_bottom + chart_height * 0.45
            return zero_y + (val / max_abs) * (chart_height * 0.45)

        zero_y = y_scale(0)
        c.setStrokeColor(colors.lightgrey); c.setLineWidth(0.8)
        c.line(chart_left, zero_y, chart_right, zero_y); c.setStrokeColor(colors.black)

        col_w = chart_width / len(bars) * 0.9
        gap   = (chart_width / len(bars)) * 0.1
        x     = chart_left + gap

        c.setFont("Helvetica", 9)
        running = beg_cash
        for name, val, kind in bars:
            if kind == "base":
                y0 = y_scale(0); y1 = y_scale(val)
            elif kind == "flow":
                y0 = y_scale(running); running += val; y1 = y_scale(running)
            else:
                y0 = y_scale(0); y1 = y_scale(val)

            bar_bottom = min(y0, y1); bar_top = max(y0, y1)
            fill = colors.Color(0.20, 0.45, 0.80) if kind == "base" else (
                   colors.Color(0.25, 0.65, 0.35) if val >= 0 else colors.Color(0.80, 0.30, 0.25))

            c.setFillColor(fill); c.setStrokeColor(colors.black)
            c.rect(x, bar_bottom, col_w, max(4, bar_top - bar_bottom), fill=1, stroke=1)

            c.setFillColor(colors.black)
            c.drawCentredString(x + col_w / 2, chart_bottom - 12, name)
            label_y = bar_top + 6 if val >= 0 else bar_bottom - 12
            c.drawCentredString(x + col_w / 2, label_y, fmt_currency(val))
            x += col_w + gap

        y = chart_bottom - 24
        hrule(y); y -= SEC_GAP

        # ---------- Assumptions (pretty) ----------
        y = section_title("Key Assumptions", y)
        pretty = [
            ("Revenue Growth YoY", "Revenue_Growth_YoY", "pct"),
            ("COGS as % of Revenue", "COGS_pct_of_Revenue", "pct"),
            ("OpEx as % of Revenue", "OpEx_pct_of_Revenue", "pct"),
            ("D&A as % of Revenue", "DA_pct_of_Revenue", "pct"),
            ("CapEx as % of Revenue", "CapEx_pct_of_Revenue", "pct"),
            ("A/R as % of Revenue", "AR_pct_of_Revenue", "pct"),
            ("Inventory as % of Revenue", "Inventory_pct_of_Revenue", "pct"),
            ("A/P as % of Revenue", "AP_pct_of_Revenue", "pct"),
            ("Tax Rate", "Tax_Rate", "pct"),
            ("Interest Rate on Debt", "Interest_Rate_on_Debt", "pct"),
            ("Debt Issuance % of Revenue", "Debt_Issuance_pct_of_Revenue", "pct"),
            ("Debt Repayment % of Revenue", "Debt_Repayment_pct_of_Revenue", "pct"),
            ("Dividend Payout % of NI", "Dividend_Payout_pct_of_NI", "pct"),
        ]
        ass_rows = []
        for label, key, kind in pretty:
            val = assumptions.get(key)
            ass_rows.append((label, fmt_pct(val) if kind == "pct" else fmt_any(val)))
        y = draw_kv_table(ass_rows, y, col_split=4.0*inch)
        y = ensure_space(y, needed=60)

        # close
        footer()
        c.showPage()
        c.save()
        return True

    except Exception as e:
        print(f"PDF export skipped (install 'reportlab' to enable). Reason: {e}")
        return False

# ==========================
# CLI + Main
# ==========================
def main():
    parser = argparse.ArgumentParser(description="3-Statement Model (2020–2029) with Assumptions & Debt Schedule")
    parser.add_argument("--query", help="Ticker or company name (e.g. AAPL or Apple)", required=False)
    parser.add_argument("--reuse-assumptions", action="store_true", help="Read overrides from existing Excel 'Assumptions' sheet if present.")
    args = parser.parse_args()

    query = args.query or input("Enter company (ticker or name, e.g., 'AAPL' or 'Apple'): ").strip()
    try:
        ticker = resolve_ticker_or_name(query)
    except Exception as e:
        print(f"Resolver: {e}")
        return

    print(f"Fetching SEC CompanyFacts for {ticker} (actuals 2020–2024) and building projections 2025–2029...")
    try:
        income_df, balance_df, cashflow_df, assumptions, debt_sched = load_financial_data(
            ticker, reuse_assumptions=args.reuse_assumptions
        )
        income_df, balance_df, cashflow_df = validate_input_data(income_df, balance_df, cashflow_df)

        # Company title (best-effort)
        try:
            maps = _load_sec_ticker_maps()
            company_title = maps[maps["ticker"].str.upper() == ticker.upper()].iloc[0]["title"]
        except Exception:
            company_title = None

        print("Linking statements…")
        income_model, balance_model, cashflow_model = create_three_statement_model(income_df, balance_df, cashflow_df)

        xlsx = f"{ticker}_three_statement_model_2020_2029.xlsx"
        print("Writing Excel (with Assumptions + Debt Schedule)…")
        save_formatted_excel(income_model, balance_model, cashflow_model, assumptions, debt_sched, xlsx, ticker, company_title)

        print("Exporting summary PDF… (optional)")
        pdf_ok = export_summary_pdf(f"{ticker}_summary.pdf", ticker, company_title, income_model, balance_model, cashflow_model, assumptions)

        print("✓ Done!")
        print(f"Excel: {xlsx}")
        if pdf_ok:
            print(f"PDF:   {ticker}_summary.pdf")

    except Exception as e:
        print(f"Error: {e}")
        print("Tip: if you edited assumptions, re-run with --reuse-assumptions to apply them.")

if __name__ == "__main__":
    main()
